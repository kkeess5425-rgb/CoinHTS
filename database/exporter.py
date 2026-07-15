"""
database/exporter.py
====================
매매 데이터 CSV / Excel 내보내기.
"""
from __future__ import annotations

import csv
import io
import json
import logging
import time
from dataclasses import asdict
from typing import Optional

logger = logging.getLogger(__name__)


class DataExporter:
    """
    매매 데이터를 CSV / Excel / JSON으로 내보내기.
    """

    # ── CSV ─────────────────────────────────────────
    def export_trades_csv(self, trades: list) -> str:
        """트레이드 기록 → CSV 문자열."""
        if not trades:
            return "데이터 없음"

        buf = io.StringIO()
        fields = [
            "ID", "심볼", "방향", "진입가", "청산가",
            "SL", "TP", "크기", "PnL(R)", "PnL(USD)",
            "진입시간", "청산시간", "보유시간(분)", "결과",
        ]
        writer = csv.writer(buf)
        writer.writerow(fields)

        for t in trades:
            entry_dt = time.strftime("%Y-%m-%d %H:%M", time.localtime(getattr(t, 'entry_ts', 0)))
            exit_dt  = time.strftime("%Y-%m-%d %H:%M", time.localtime(getattr(t, 'exit_ts', 0)))
            hold_min = round((getattr(t,'exit_ts',0) - getattr(t,'entry_ts',0)) / 60, 1)
            result   = "WIN" if getattr(t,'pnl_r',0) > 0 else ("LOSS" if getattr(t,'pnl_r',0) < 0 else "BE")
            writer.writerow([
                getattr(t, 'id', ''),
                getattr(t, 'symbol', ''),
                getattr(t, 'direction', ''),
                getattr(t, 'entry', ''),
                getattr(t, 'exit_price', ''),
                getattr(t, 'sl', ''),
                getattr(t, 'tp', ''),
                getattr(t, 'size', ''),
                getattr(t, 'pnl_r', ''),
                getattr(t, 'pnl_usd', ''),
                entry_dt, exit_dt, hold_min, result,
            ])
        return buf.getvalue()

    def export_journal_csv(self, entries: list) -> str:
        """매매일지 → CSV."""
        if not entries:
            return "데이터 없음"
        buf    = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow([
            "ID", "심볼", "방향", "진입가", "청산가", "SL", "TP",
            "PnL(R)", "PnL(USD)", "진입시간", "청산이유", "진입근거", "실수",
        ])
        for e in entries:
            ts     = time.strftime("%Y-%m-%d %H:%M", time.localtime(getattr(e,'entry_ts',0)))
            mistakes = ", ".join(m.kind for m in getattr(e,'mistakes',[]))
            writer.writerow([
                getattr(e,'id',''), getattr(e,'symbol',''),
                getattr(e,'direction',''), getattr(e,'entry',''),
                getattr(e,'exit_price',''), getattr(e,'sl',''), getattr(e,'tp',''),
                getattr(e,'pnl_r',''), getattr(e,'pnl_usd',''),
                ts, getattr(e,'exit_reason',''), getattr(e,'entry_reason',''), mistakes,
            ])
        return buf.getvalue()

    def export_stats_json(self, stats) -> str:
        """통계 → JSON."""
        try:
            from dataclasses import asdict as _asdict
            return json.dumps(_asdict(stats), ensure_ascii=False, indent=2)
        except Exception:
            return json.dumps(vars(stats), ensure_ascii=False, indent=2, default=str)

    def export_excel(self, trades: list, entries: list, stats=None) -> bytes:
        """Excel 내보내기 (openpyxl 필요)."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError("pip install openpyxl 필요")

        wb = openpyxl.Workbook()

        # ── 트레이드 시트 ──
        ws = wb.active
        ws.title = "트레이드"
        headers = ["ID", "심볼", "방향", "진입가", "청산가", "SL", "TP",
                   "PnL(R)", "PnL(USD)", "진입시간", "청산시간", "결과"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font  = Font(bold=True, color="C9D1D9")
            cell.fill  = PatternFill("solid", fgColor="161B22")
            cell.alignment = Alignment(horizontal="center")

        for row, t in enumerate(trades, 2):
            pnl_r   = getattr(t, 'pnl_r', 0)
            result  = "WIN" if pnl_r > 0 else ("LOSS" if pnl_r < 0 else "BE")
            fill_color = "0D2B1A" if result == "WIN" else ("2D1117" if result == "LOSS" else "21262D")
            vals = [
                getattr(t,'id',''), getattr(t,'symbol',''), getattr(t,'direction',''),
                getattr(t,'entry',0), getattr(t,'exit_price',0),
                getattr(t,'sl',0), getattr(t,'tp',0),
                pnl_r, getattr(t,'pnl_usd',0),
                time.strftime("%Y-%m-%d %H:%M", time.localtime(getattr(t,'entry_ts',0))),
                time.strftime("%Y-%m-%d %H:%M", time.localtime(getattr(t,'exit_ts',0))),
                result,
            ]
            for col, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=col, value=v)
                cell.fill = PatternFill("solid", fgColor=fill_color)

        # 컬럼 너비 자동 조정
        for col in ws.columns:
            max_len = max(len(str(c.value or '')) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max(12, min(max_len+2, 30))

        # ── 통계 시트 ──
        if stats:
            ws2 = wb.create_sheet("통계")
            stat_items = [
                ("총 트레이드",  getattr(stats, 'total_trades', 0)),
                ("승률 (%)",     getattr(stats, 'win_rate', 0)),
                ("Profit Factor",getattr(stats, 'profit_factor', 0)),
                ("Sharpe Ratio", getattr(stats, 'sharpe_ratio', 0)),
                ("Sortino Ratio",getattr(stats, 'sortino_ratio', 0)),
                ("MDD (%)",      getattr(stats, 'max_drawdown', 0)),
                ("기대값 (R)",   getattr(stats, 'expectancy', 0)),
                ("누적 R",       getattr(stats, 'total_r', 0)),
                ("평균 보유(분)",getattr(stats, 'avg_hold_minutes', 0)),
            ]
            for row, (k, v) in enumerate(stat_items, 1):
                ws2.cell(row=row, column=1, value=k).font = Font(bold=True, color="8B949E")
                ws2.cell(row=row, column=2, value=round(float(v), 3) if v else 0)
            ws2.column_dimensions['A'].width = 18
            ws2.column_dimensions['B'].width = 12

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
